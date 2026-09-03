import http.server
import json
import os
import re
import sys
import urllib.parse
from http.server import HTTPServer, SimpleHTTPRequestHandler
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure UTF-8 output
sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PORT = 8080
DEFAULT_EXCEL_PATH = r"C:\Users\500744\OneDrive - pea.co.th\Desktop\ข้อมูลคำนวณน้ำหนักอุปกรณ์(งานรื้อถอน).xlsx"
DEFAULT_PDF_PATH = r"C:\Users\500744\OneDrive - pea.co.th\Desktop\018.pdf"
MASTER_JSON_PATH = os.path.join(BASE_DIR, "master_data.json")

def get_available_desktop_files():
    """Find all available 018 PDF or Excel files on Desktop and current directory"""
    search_dirs = [
        os.path.join(BASE_DIR, "samples"),
        r"C:\Users\500744\OneDrive - pea.co.th\Desktop",
        r"C:\Users\500744\Desktop",
        os.path.join(os.path.expanduser("~"), "OneDrive - pea.co.th", "Desktop"),
        os.path.join(os.path.expanduser("~"), "Desktop"),
        BASE_DIR
    ]
    seen_paths = set()
    found = []
    for sdir in search_dirs:
        if not os.path.exists(sdir):
            continue
        try:
            for fname in os.listdir(sdir):
                if fname.startswith("~$"):
                    continue
                lower = fname.lower()
                if ("018" in lower or "zpsr" in lower or "รายงาน" in lower) and (lower.endswith(".pdf") or lower.endswith(".xlsx")):
                    full_p = os.path.join(sdir, fname)
                    if full_p not in seen_paths and os.path.isfile(full_p):
                        seen_paths.add(full_p)
                        size_kb = round(os.path.getsize(full_p) / 1024, 1)
                        # Short label for UI
                        label = fname
                        if "กำแพงแสน" in fname:
                            label = "018-กำแพงแสน.pdf (งานรื้อถอน)"
                        elif "PED" in fname:
                            label = "018-PED0001201261_1.pdf (งานสถานีไฟฟ้า)"
                        elif fname == "018.pdf":
                            label = "018.pdf (ตัวอย่างมาตรฐาน)"
                        found.append({
                            'filename': fname,
                            'path': full_p,
                            'size_kb': size_kb,
                            'label': label
                        })
        except Exception as e:
            pass

    # Sort so that 018.pdf is first, followed by others
    found.sort(key=lambda x: (0 if x['filename'] == '018.pdf' else 1, x['filename']))
    return found

def load_master_data():
    """Load equipment weight master data from JSON or fallback to Excel"""
    if os.path.exists(MASTER_JSON_PATH):
        try:
            with open(MASTER_JSON_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print("Error loading master_data.json:", e)

    # Fallback to loading directly from Excel
    if os.path.exists(DEFAULT_EXCEL_PATH):
        items = []
        try:
            wb = openpyxl.load_workbook(DEFAULT_EXCEL_PATH, data_only=True)
            sheet = wb['DATA']
            for r in range(3, sheet.max_row + 1):
                vals = [sheet.cell(r, c).value for c in range(1, 7)]
                code_raw = str(vals[0] or "").strip()
                name = str(vals[1] or "").strip()
                wt = vals[2]
                unit = str(vals[5] or "").strip()
                if not name and not code_raw:
                    continue
                code_clean = re.sub(r'\D', '', code_raw)
                try:
                    wt_val = float(wt) if wt is not None else 0.0
                except:
                    wt_val = 0.0
                formatted_code = ""
                if len(code_clean) == 10:
                    formatted_code = f"{code_clean[0]}-{code_clean[1:3]}-{code_clean[3:6]}-{code_clean[6:]}"
                items.append({
                    'id': r - 2,
                    'code': code_clean,
                    'formatted_code': formatted_code,
                    'raw_code': code_raw,
                    'name': name,
                    'weight_per_unit': wt_val,
                    'unit': unit or "กก."
                })
            with open(MASTER_JSON_PATH, "w", encoding="utf-8") as f:
                json.dump(items, f, ensure_ascii=False, indent=2)
            return items
        except Exception as e:
            print("Error loading master from Excel:", e)
    return []

def parse_sap_excel(file_path_or_bytes):
    """
    Parse PEA SAP Material / Closeout Report from Excel (.xlsx / .xls).
    Extracts metadata and table rows.
    """
    import io
    if isinstance(file_path_or_bytes, bytes):
        wb = openpyxl.load_workbook(io.BytesIO(file_path_or_bytes), data_only=True)
    else:
        wb = openpyxl.load_workbook(file_path_or_bytes, data_only=True)

    sheet = wb.active
    metadata = {
        'job_no': '',
        'job_name': '',
        'person_name': '',
        'person_id': '',
        'cost_center_req': '',
        'cost_center_resp': '',
        'print_date': '',
        'total_report_pages': 1
    }

    # 1. Scan first 20 rows for metadata and column header row
    header_row_idx = None
    col_map = {
        'code': None,
        'desc': None,
        'unit': None,
        'est': None,
        'issued': None,
        'good_return': None,
        'dam_return': None,
        'installed': None,
        'section': None
    }

    for r in range(1, min(25, sheet.max_row + 1)):
        row_vals = [str(sheet.cell(r, c).value or "").strip() for c in range(1, min(30, sheet.max_column + 1))]
        row_text = " ".join(row_vals)

        # Check metadata
        if not metadata['job_no']:
            m = re.search(r'(?:หมายเลขงาน|เลขที่งาน|งานเลขที่|Job\s*No)\s*[:\s]*([A-Z0-9\.\-]+)', row_text) or re.search(r'([IP]-\d{2}-[A-Z0-9\.\-]+)', row_text) or re.search(r'([IP]-[A-Z0-9\.\-]+)', row_text)
            if m: metadata['job_no'] = m.group(1).strip()

        if not metadata['job_name']:
            m = re.search(r'(?:ชื่องาน|องาน|Project)\s*[:\s]*([^_\n\|]+)', row_text)
            if m: metadata['job_name'] = m.group(1).strip()

        if not metadata['person_name']:
            m = re.search(r'(?:นาย|นางสาว|นาง)\s*([^\s]+)\s+([^\s]+)', row_text)
            if m: metadata['person_name'] = m.group(0)

        # Check if this row looks like the table header
        has_code_col = any(('รหัส' in v or 'พัสดุ' in v or 'material' in v.lower() or 'code' in v.lower()) for v in row_vals)
        has_qty_or_desc = any(('รายการ' in v or 'จำนวน' in v or 'ประมาณ' in v or 'desc' in v.lower() or 'หน่วย' in v) for v in row_vals)

        if has_code_col and has_qty_or_desc and header_row_idx is None:
            header_row_idx = r
            for c_idx, val in enumerate(row_vals, start=1):
                v_low = val.lower()
                if ('รหัส' in val or 'code' in v_low or 'material' in v_low) and col_map['code'] is None:
                    col_map['code'] = c_idx
                elif ('รายการ' in val or 'รายละเอียด' in val or 'desc' in v_low or 'name' in v_low) and col_map['desc'] is None:
                    col_map['desc'] = c_idx
                elif ('หน่วย' in val or 'unit' in v_low) and col_map['unit'] is None:
                    col_map['unit'] = c_idx
                elif ('ประมาณ' in val or 'est' in v_low) and col_map['est'] is None:
                    col_map['est'] = c_idx
                elif ('เบิก' in val or 'issue' in v_low) and col_map['issued'] is None:
                    col_map['issued'] = c_idx
                elif ('ดี' in val or 'good' in v_low) and col_map['good_return'] is None:
                    col_map['good_return'] = c_idx
                elif ('ชำรุด' in val or 'dam' in v_low) and col_map['dam_return'] is None:
                    col_map['dam_return'] = c_idx
                elif ('ติดตั้ง' in val or 'install' in v_low) and col_map['installed'] is None:
                    col_map['installed'] = c_idx
                elif ('แผนก' in val or 'ตอน' in val or 'sec' in v_low) and col_map['section'] is None:
                    col_map['section'] = c_idx

    # If no header found, fallback
    if header_row_idx is None:
        header_row_idx = 1
        col_map['code'] = 1
        col_map['desc'] = 2
        col_map['unit'] = 3
        col_map['est'] = 4

    if col_map['desc'] is None and col_map['code'] is not None:
        col_map['desc'] = col_map['code'] + 1

    raw_items = []
    current_section = "งานทั่วไป"
    seen_codes = set()

    for r in range(header_row_idx + 1, sheet.max_row + 1):
        def get_val(c):
            if not c: return ""
            return sheet.cell(r, c).value

        code_raw = str(get_val(col_map['code']) or "").strip()
        desc_raw = str(get_val(col_map['desc']) or "").strip()

        # Section row
        if not code_raw and desc_raw:
            if any(k in desc_raw for k in ['แผนก', 'ตอนที่', 'TL-', 'HT-', 'TR-', 'รื้อถอน', 'ก่อสร้าง']):
                current_section = desc_raw
                continue

        code_clean = re.sub(r'\D', '', code_raw)
        if not code_clean or len(code_clean) < 7:
            continue

        formatted_code = code_raw
        if len(code_clean) == 10:
            formatted_code = f"{code_clean[0]}-{code_clean[1:3]}-{code_clean[3:6]}-{code_clean[6:]}"

        def get_float(c):
            v = get_val(c)
            if v is None: return 0.0
            try: return float(str(v).replace(',', ''))
            except: return 0.0

        unit = str(get_val(col_map['unit']) or "EA").strip()
        qty_est = get_float(col_map['est'])
        qty_issued = get_float(col_map['issued'])
        qty_good = get_float(col_map['good_return'])
        qty_dam = get_float(col_map['dam_return'])
        qty_installed = get_float(col_map['installed'])

        sec = str(get_val(col_map['section']) or "").strip() or current_section
        is_dismantle = ("-R-E" in sec) or ("รื้อถอน" in sec) or ("dismantle" in sec.lower())
        calc_qty = qty_est if is_dismantle else (qty_good + qty_dam if (qty_good or qty_dam) else (qty_installed or qty_issued))

        item_key = f"{code_clean}_{desc_raw[:15]}"
        if item_key not in seen_codes:
            seen_codes.add(item_key)
            raw_items.append({
                'page': 1,
                'section': sec,
                'is_dismantle': is_dismantle,
                'code': formatted_code,
                'code_10': code_clean,
                'desc': desc_raw,
                'unit': unit,
                'qty_estimate': qty_est,
                'qty_issued': qty_issued,
                'qty_good_return': qty_good,
                'qty_damaged_return': qty_dam,
                'qty_installed': qty_installed,
                'cost_issued': 0.0,
                'cost_return': 0.0,
                'suggested_qty': calc_qty
            })

    return {
        'metadata': metadata,
        'items': raw_items
    }

def parse_sap_pdf(file_path_or_bytes):
    """
    Parse PEA SAP Construction Closeout Report (ZPSR018) from PDF.
    Extracts project header and table rows categorized by section.
    Includes fast validation to reject non-SAP documents without hanging.
    """
    from pdfminer.high_level import extract_pages
    from pdfminer.layout import LTTextContainer
    import io

    # 1. Fast Validation Check (Inspect first 2 pages in <2s)
    if isinstance(file_path_or_bytes, bytes):
        fp_check = io.BytesIO(file_path_or_bytes)
        first_pages = list(extract_pages(fp_check, maxpages=2))
    else:
        first_pages = list(extract_pages(file_path_or_bytes, maxpages=2))

    first_pages_text = ""
    for p in first_pages:
        for el in p:
            if isinstance(el, LTTextContainer):
                first_pages_text += el.get_text() + " "

    # Look for SAP material code patterns like 1-xx-xxx-xxxx
    has_sap_material_codes = bool(re.search(r'[19]-\d{2}-\d{3}-\d{4}', first_pages_text))
    if not has_sap_material_codes:
        raise ValueError(
            "ไม่พบตารางรหัสพัสดุ (เช่น 1-xx-xxx-xxxx) ในเอกสารนี้ "
            "ไฟล์ที่เลือกอาจเป็นเอกสารบันทึกข้อความหรือสไลด์บรรยาย ไม่ใช่รายงานสรุปพัสดุ ZPSR018 จาก SAP "
            "กรุณาตรวจสอบและอัปโหลดไฟล์รายงานผลพัสดุ ZPSR018 (PDF หรือ Excel) จากระบบ SAP"
        )

    # 2. Extract full pages
    if isinstance(file_path_or_bytes, bytes):
        fp = io.BytesIO(file_path_or_bytes)
        pages = list(extract_pages(fp))
    else:
        pages = list(extract_pages(file_path_or_bytes))

    # Detect page cycle (e.g. 3 copies in a 21-page PDF: look for "หน้า 1 / X")
    num_pages_to_process = len(pages)
    for p_idx, p in enumerate(pages[:8]):
        text_sample = ""
        for el in p:
            if isinstance(el, LTTextContainer):
                text_sample += el.get_text() + " "
        m_total = re.search(r'หน้า\s*1\s*/\s*(\d+)', text_sample)
        if m_total:
            total_report_pages = int(m_total.group(1))
            num_pages_to_process = min(len(pages), total_report_pages)
            break

    metadata = {
        'job_no': '',
        'job_name': '',
        'person_name': '',
        'person_id': '',
        'cost_center_req': '',
        'cost_center_resp': '',
        'print_date': '',
        'total_report_pages': num_pages_to_process
    }

    # Extract clean metadata from the first few pages
    full_sample_text = ""
    for p in pages[:min(6, len(pages))]:
        for el in p:
            if isinstance(el, LTTextContainer):
                full_sample_text += el.get_text() + " "
    norm_text = " ".join(full_sample_text.split())

    # Detect if whole document is a dedicated dismantle report
    # (e.g. "(ปิดงานรื้อถอน)", "รื้อถอน", "ถอน", "(cid:202)")
    doc_is_dismantle = bool(re.search(r'(?:ปิดงานรื้อถอน|รื้อถอน|\u0e16\u0e2d\u0e19|จำนวนพัสดุที่รื้อถอน|\(cid:202\).*?\u0e16\u0e2d\u0e19)', full_sample_text))

    # Person
    m_person = re.search(r'นาย([^\s]+)\s+([^\s]+)\s+รหัสประจําตัว\s+(\d+)', norm_text)
    if m_person:
        metadata['person_name'] = f"นาย{m_person.group(1)} {m_person.group(2)}"
        metadata['person_id'] = m_person.group(3)
    else:
        m_p2 = re.search(r'(?:นาย|นางสาว|นาง)\s*([^\s]+)\s+([^\s]+)', norm_text)
        if m_p2: metadata['person_name'] = m_p2.group(0)
        m_pid = re.search(r'รหัสประจําตัว\s*(\d+)', norm_text) or re.search(r'รหัสประจำตัว\s*(\d+)', norm_text)
        if m_pid: metadata['person_id'] = m_pid.group(1)

    # Job Name & No
    m_job = re.search(r'(?:ชื่องาน|องาน)\s+(.*?)\s+(?:หมายเลขงาน|เลขที่งาน)\s+([A-Z0-9\.\-]+)', norm_text)
    if m_job:
        metadata['job_name'] = m_job.group(1).strip()
        metadata['job_no'] = m_job.group(2).strip()

    if not metadata['job_no']:
        m_job_no = (
            re.search(r'(?:หมายเลขงาน|เลขที่งาน|งานเลขที่)\s*([A-Z0-9\.\-]+)', norm_text)
            or re.search(r'([IP]-\d{2}-[A-Z0-9\.\-]+)', norm_text)
            or re.search(r'([IP]-[A-Z0-9\.\-]+)', norm_text)
        )
        if m_job_no: metadata['job_no'] = m_job_no.group(1).strip()

    if not metadata['job_name']:
        m_jn = re.search(r'(?:ชื่องาน|องาน)\s+([^_\n\|]{3,50})', norm_text)
        if m_jn: metadata['job_name'] = m_jn.group(1).strip()

    m_cost = re.search(r'รับผิดชอบ\s+([A-Z0-9]+)', norm_text)
    if m_cost:
        metadata['cost_center_resp'] = m_cost.group(1)

    m_date = re.search(r'พิมพ์\s+(\d{2}\.\d{2}\.\d{4})', norm_text) or re.search(r'(\d{2}\.\d{2}\.\d{4})', norm_text)
    if m_date:
        metadata['print_date'] = m_date.group(1)

    raw_items = []
    current_section = "งานรื้อถอน (ปิดงานรื้อถอน)" if doc_is_dismantle else "งานทั่วไป"
    seen_keys = set()

    for page_idx in range(num_pages_to_process):
        page = pages[page_idx]
        elements = []
        for el in page:
            if isinstance(el, LTTextContainer):
                for line in el:
                    txt = line.get_text().strip()
                    if txt:
                        elements.append({
                            'y0': line.y0, 'y1': line.y1,
                            'x0': line.x0, 'x1': line.x1,
                            'text': txt
                        })
        elements.sort(key=lambda x: -x['y1'])

        # Group by vertical line position (tolerance 3.5 points)
        line_groups = []
        for el in elements:
            matched = False
            for grp in line_groups:
                if abs(grp[0]['y1'] - el['y1']) <= 3.5:
                    grp.append(el)
                    matched = True
                    break
            if not matched:
                line_groups.append([el])

        for grp in line_groups:
            grp.sort(key=lambda x: x['x0'])
            full_line_text = " ".join([x['text'] for x in grp])

            # Check section header (e.g. 1. 7000656656 / TL-R-E แผนกรื้อถอนสายส่งภายนอก...)
            m_sec = re.search(r'(\d+\.\s+\d+\s+/\s+([A-Z0-9\-]+)\s+[^_\|]+)', full_line_text)
            if m_sec:
                current_section = m_sec.group(1).strip()
                continue

            # Check material item code (1-xx-xxx-xxxx or 9-xx-xxx-xxxx)
            m_code = re.search(r'([19]-\d{2}-\d{3}-\d{4})', full_line_text)
            if m_code:
                code_str = m_code.group(1)
                code_clean = re.sub(r'\D', '', code_str)

                desc = ""
                unit = "EA"
                qty_estimate = 0.0
                qty_issued = 0.0
                qty_good_return = 0.0
                qty_damaged_return = 0.0
                qty_installed = 0.0
                qty_dismantled_record = 0.0
                cost_issued = 0.0
                cost_return = 0.0

                for it in grp:
                    x = it['x0']
                    val = it['text'].replace('|', '').strip()
                    if not val:
                        continue
                    if x < 220:
                        cleaned_val = re.sub(r'^[19]-\d{2}-\d{3}-\d{4}\s*(NA|NS|SL|NN)?\s*', '', val).strip()
                        if cleaned_val:
                            desc += " " + cleaned_val
                    elif 220 <= x < 255:
                        unit = val
                    elif 255 <= x < 320:
                        try: qty_estimate = float(val.replace(',', ''))
                        except: pass
                    elif 320 <= x < 380:
                        try: qty_issued = float(val.replace(',', ''))
                        except: pass
                    elif 380 <= x < 440:
                        try: qty_good_return = float(val.replace(',', ''))
                        except: pass
                    elif 440 <= x < 505:
                        try: qty_damaged_return = float(val.replace(',', ''))
                        except: pass
                    elif 505 <= x < 565:
                        try: qty_installed = float(val.replace(',', ''))
                        except: pass
                    elif 565 <= x < 635:
                        try: cost_issued = float(val.replace(',', ''))
                        except: pass
                    elif 635 <= x < 700:
                        try: cost_return = float(val.replace(',', ''))
                        except: pass
                    elif 700 <= x < 830:
                        try: qty_dismantled_record = float(val.replace(',', ''))
                        except: pass

                desc = desc.strip()
                # Determine department type
                is_dismantle = doc_is_dismantle or ("-R-E" in current_section) or ("รื้อถอน" in current_section) or ("\u0e16\u0e2d\u0e19" in current_section)

                # Default suggested calculation quantity
                if is_dismantle:
                    if qty_dismantled_record > 0:
                        calc_qty = qty_dismantled_record
                    elif qty_estimate > 0:
                        calc_qty = qty_estimate
                    elif (qty_good_return + qty_damaged_return) > 0:
                        calc_qty = qty_good_return + qty_damaged_return
                    else:
                        calc_qty = qty_installed
                else:
                    calc_qty = (qty_good_return + qty_damaged_return) if (qty_good_return or qty_damaged_return) else qty_installed

                item_key = f"{page_idx+1}_{code_str}_{desc[:15]}"
                if item_key not in seen_keys:
                    seen_keys.add(item_key)
                    raw_items.append({
                        'page': page_idx + 1,
                        'section': current_section,
                        'is_dismantle': is_dismantle,
                        'code': code_str,
                        'code_10': code_clean,
                        'desc': desc,
                        'unit': unit,
                        'qty_estimate': qty_estimate,
                        'qty_issued': qty_issued,
                        'qty_good_return': qty_good_return,
                        'qty_damaged_return': qty_damaged_return,
                        'qty_installed': qty_installed,
                        'qty_dismantled_record': qty_dismantled_record,
                        'cost_issued': cost_issued,
                        'cost_return': cost_return,
                        'suggested_qty': calc_qty
                    })

    return {
        'metadata': metadata,
        'items': raw_items
    }

def match_items_with_master(pdf_items, master_data):
    """
    Smart Matching Engine:
    Match PDF items with master equipment scrap weights.
    Returns matched items with match_type: 'exact', 'suggested', or 'none'.
    """
    master_by_10 = {}
    master_by_6 = {}
    for m in master_data:
        code_10 = m.get('code', '')
        if code_10:
            master_by_10[code_10] = m
            if len(code_10) >= 6:
                prefix = code_10[:6]
                if prefix not in master_by_6:
                    master_by_6[prefix] = m

    enriched_items = []
    for item in pdf_items:
        code_10 = item.get('code_10', '')
        desc = item.get('desc', '')
        matched_master = None
        match_type = 'none'

        # 1. Exact 10-digit code match
        if code_10 in master_by_10:
            matched_master = master_by_10[code_10]
            match_type = 'exact'
        elif len(code_10) >= 6 and code_10[:6] in master_by_6:
            # 2. Category prefix match
            matched_master = master_by_6[code_10[:6]]
            match_type = 'suggested'
        else:
            # 3. Fuzzy keyword match on description
            desc_lower = desc.lower()
            best_match = None
            for m in master_data:
                m_name = m.get('name', '').lower()
                # Keyword checks
                if 'bolt' in desc_lower and 'สลักเกลียว' in m_name:
                    best_match = m
                    break
                elif 'thimble' in desc_lower and 'ทิมเบิล' in m_name:
                    best_match = m
                    break
                elif 'crossarm' in desc_lower and 'ประกับ' in m_name:
                    best_match = m
                    break
                elif 'insulator' in desc_lower and 'ลูกถ้วย' in m_name:
                    best_match = m
                    break
                elif 'rack' in desc_lower and 'แร็ค' in m_name:
                    best_match = m
                    break
                elif 'channel' in desc_lower and 'รางน้ำ' in m_name:
                    best_match = m
                    break
                elif 'bracket' in desc_lower and 'เหล็กคอน' in m_name:
                    best_match = m
                    break
            if best_match:
                matched_master = best_match
                match_type = 'suggested'

        weight_per_unit = matched_master['weight_per_unit'] if matched_master else 0.0
        master_name = matched_master['name'] if matched_master else ''
        calc_qty = item.get('suggested_qty', 0.0)

        # By default, items in dismantle section (-R-E) with matched scrap weight are selected
        is_selected = item.get('is_dismantle', False) and (weight_per_unit > 0 or match_type in ['exact', 'suggested'])

        enriched_items.append({
            **item,
            'match_type': match_type,
            'master_name': master_name,
            'master_code': matched_master.get('code', '') if matched_master else '',
            'weight_per_unit': weight_per_unit,
            'calc_qty': calc_qty,
            'total_weight': round(calc_qty * weight_per_unit, 3),
            'selected': is_selected
        })

    return enriched_items

def clean_excel_str(val):
    """Remove control characters that openpyxl rejects"""
    if val is None:
        return ""
    if isinstance(val, str):
        return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', val).strip()
    return val

def generate_export_excel(data):
    """
    Generate official-looking PEA Scrap Return Excel report (.xlsx)
    using openpyxl with purple/orange PEA branding and formulas.
    """
    metadata = data.get('metadata', {})
    items = data.get('items', [])
    summary = data.get('summary', {})

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "การคืนเศษเหล็ก"

    # Show gridlines
    ws.views.sheetView[0].showGridLines = True

    # Color Palette - PEA Branding
    purple_fill = PatternFill(start_color="5E136E", end_color="5E136E", fill_type="solid")
    purple_light_fill = PatternFill(start_color="F2E6F5", end_color="F2E6F5", fill_type="solid")
    orange_fill = PatternFill(start_color="F26522", end_color="F26522", fill_type="solid")
    header_fill = PatternFill(start_color="6C1D7E", end_color="6C1D7E", fill_type="solid")
    gray_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    total_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

    font_title = Font(name="Prompt", size=16, bold=True, color="FFFFFF")
    font_sub = Font(name="Prompt", size=11, color="FFFFFF")
    font_header = Font(name="Prompt", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Sarabun", size=10)
    font_bold = Font(name="Prompt", size=11, bold=True)
    font_total = Font(name="Prompt", size=12, bold=True, color="D84315")

    thin_border = Border(
        left=Side(style='thin', color='D0D5DD'),
        right=Side(style='thin', color='D0D5DD'),
        top=Side(style='thin', color='D0D5DD'),
        bottom=Side(style='thin', color='D0D5DD')
    )
    double_bottom_border = Border(
        left=Side(style='thin', color='D0D5DD'),
        right=Side(style='thin', color='D0D5DD'),
        top=Side(style='thin', color='D0D5DD'),
        bottom=Side(style='double', color='5E136E')
    )

    # 1. Main Title Banner
    ws.merge_cells("A1:H1")
    ws["A1"] = "การไฟฟ้าส่วนภูมิภาค (PROVINCIAL ELECTRICITY AUTHORITY)"
    ws["A1"].font = font_title
    ws["A1"].fill = purple_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:H2")
    ws["A2"] = "รายงานการคำนวณการคืนพัสดุประเภทเศษเหล็ก (งานก่อสร้างและรื้อถอนระบบไฟฟ้า)"
    ws["A2"].font = font_sub
    ws["A2"].fill = purple_fill
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    # 2. Metadata Block
    meta_rows = [
        ("หมายเลขงาน:", clean_excel_str(metadata.get('job_no', '-')), "วันที่รายงาน:", clean_excel_str(metadata.get('print_date', '-'))),
        ("ชื่องาน:", clean_excel_str(metadata.get('job_name', '-')), "ศูนย์ต้นทุน:", clean_excel_str(metadata.get('cost_center_resp', '-'))),
        ("ผู้ควบคุมงาน:", clean_excel_str(metadata.get('person_name', '-')), "รหัสพนักงาน:", clean_excel_str(metadata.get('person_id', '-')))
    ]

    curr_r = 4
    for r_idx, (k1, v1, k2, v2) in enumerate(meta_rows):
        row_num = curr_r + r_idx
        ws.cell(row=row_num, column=1, value=k1).font = font_bold
        ws.cell(row=row_num, column=2, value=v1).font = font_data
        ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=4)
        ws.cell(row=row_num, column=5, value=k2).font = font_bold
        ws.cell(row=row_num, column=6, value=v2).font = font_data
        ws.merge_cells(start_row=row_num, start_column=6, end_row=row_num, end_column=8)
        ws.row_dimensions[row_num].height = 20

    # 3. Table Headers
    headers = [
        ("ลำดับ", 6, Alignment(horizontal="center", vertical="center")),
        ("รหัสพัสดุ SAP", 16, Alignment(horizontal="center", vertical="center")),
        ("รายการพัสดุ (SAP / รายละเอียด)", 36, Alignment(horizontal="left", vertical="center")),
        ("รายการอ้างอิงฐานข้อมูล", 30, Alignment(horizontal="left", vertical="center")),
        ("หน่วย", 8, Alignment(horizontal="center", vertical="center")),
        ("จำนวน", 12, Alignment(horizontal="right", vertical="center")),
        ("กก./หน่วย", 12, Alignment(horizontal="right", vertical="center")),
        ("น้ำหนักรวม (กก.)", 18, Alignment(horizontal="right", vertical="center"))
    ]

    header_row = 8
    ws.row_dimensions[header_row].height = 26
    for c_idx, (h_name, col_w, align) in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c_idx, value=h_name)
        cell.font = font_header
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        col_letter = get_column_letter(c_idx)
        ws.column_dimensions[col_letter].width = col_w

    # 4. Insert Items
    start_item_row = 9
    current_row = start_item_row
    item_index = 1

    for it in items:
        # Check if selected
        if not it.get('selected', True):
            continue

        r = current_row
        ws.row_dimensions[r].height = 22
        is_even = (item_index % 2 == 0)
        row_fill = gray_zebra if is_even else PatternFill(fill_type=None)

        qty_val = float(it.get('calc_qty', it.get('qty', 0)))
        wt_val = float(it.get('weight_per_unit', 0))

        # Col 1: Index
        c1 = ws.cell(row=r, column=1, value=item_index)
        c1.alignment = Alignment(horizontal="center", vertical="center")

        # Col 2: Code
        c2 = ws.cell(row=r, column=2, value=clean_excel_str(it.get('code', '')))
        c2.alignment = Alignment(horizontal="center", vertical="center")

        # Col 3: Desc
        c3 = ws.cell(row=r, column=3, value=clean_excel_str(it.get('desc', '')))
        c3.alignment = Alignment(horizontal="left", vertical="center")

        # Col 4: Master Name
        c4 = ws.cell(row=r, column=4, value=clean_excel_str(it.get('master_name', '') or it.get('desc', '')))
        c4.alignment = Alignment(horizontal="left", vertical="center")

        # Col 5: Unit
        c5 = ws.cell(row=r, column=5, value=clean_excel_str(it.get('unit', '')))
        c5.alignment = Alignment(horizontal="center", vertical="center")

        # Col 6: Qty
        c6 = ws.cell(row=r, column=6, value=qty_val)
        c6.number_format = "#,##0.00"
        c6.alignment = Alignment(horizontal="right", vertical="center")

        # Col 7: Weight per unit
        c7 = ws.cell(row=r, column=7, value=wt_val)
        c7.number_format = "#,##0.00"
        c7.alignment = Alignment(horizontal="right", vertical="center")

        # Col 8: Formula Total Weight = Qty * Weight
        c8 = ws.cell(row=r, column=8, value=f"=F{r}*G{r}")
        c8.number_format = "#,##0.00"
        c8.alignment = Alignment(horizontal="right", vertical="center")

        for col_idx in range(1, 9):
            ws.cell(row=r, column=col_idx).border = thin_border
            if row_fill.fill_type:
                ws.cell(row=r, column=col_idx).fill = row_fill
            ws.cell(row=r, column=col_idx).font = font_data

        current_row += 1
        item_index += 1

    # If no items were selected
    if current_row == start_item_row:
        ws.merge_cells(start_row=start_item_row, start_column=1, end_row=start_item_row, end_column=8)
        ws.cell(row=start_item_row, column=1, value="ไม่มีรายการพัสดุที่เลือกคำนวณ").alignment = Alignment(horizontal="center")
        current_row += 1

    # 5. Grand Total Row
    total_row = current_row
    ws.row_dimensions[total_row].height = 28
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=7)
    tot_label = ws.cell(row=total_row, column=1, value="น้ำหนักเศษเหล็กรวมทั้งสิ้น (กิโลกรัม / kg):")
    tot_label.font = font_bold
    tot_label.alignment = Alignment(horizontal="right", vertical="center")

    tot_val = ws.cell(row=total_row, column=8, value=f"=SUM(H{start_item_row}:H{total_row-1})")
    tot_val.font = font_total
    tot_val.number_format = "#,##0.00"
    tot_val.alignment = Alignment(horizontal="right", vertical="center")

    for col_idx in range(1, 9):
        ws.cell(row=total_row, column=col_idx).fill = total_fill
        ws.cell(row=total_row, column=col_idx).border = double_bottom_border

    # Tons row
    tons_row = total_row + 1
    ws.row_dimensions[tons_row].height = 24
    ws.merge_cells(start_row=tons_row, start_column=1, end_row=tons_row, end_column=7)
    tons_label = ws.cell(row=tons_row, column=1, value="น้ำหนักเศษเหล็กรวม (ตัน / Metric Tons):")
    tons_label.font = font_bold
    tons_label.alignment = Alignment(horizontal="right", vertical="center")

    tons_val = ws.cell(row=tons_row, column=8, value=f"=H{total_row}/1000")
    tons_val.font = font_total
    tons_val.number_format = "#,##0.000"
    tons_val.alignment = Alignment(horizontal="right", vertical="center")

    for col_idx in range(1, 9):
        ws.cell(row=tons_row, column=col_idx).fill = total_fill
        ws.cell(row=tons_row, column=col_idx).border = thin_border

    # 6. Sign-off Blocks for PEA
    sign_start_row = tons_row + 3
    sign_titles = [
        ("ลงชื่อ........................................................", "ผู้ส่งคืน / ผู้ควบคุมงาน"),
        ("ลงชื่อ........................................................", "ผู้ตรวจสอบ / คณะกรรมการ"),
        ("ลงชื่อ........................................................", "ผู้อนุมัติ / หัวหน้าแผนก")
    ]

    for s_idx, (line1, line2) in enumerate(sign_titles):
        col_st = 1 + (s_idx * 3)
        col_ed = col_st + 1
        ws.merge_cells(start_row=sign_start_row, start_column=col_st, end_row=sign_start_row, end_column=col_ed)
        c_sig1 = ws.cell(row=sign_start_row, column=col_st, value=line1)
        c_sig1.alignment = Alignment(horizontal="center")
        c_sig1.font = font_data

        ws.merge_cells(start_row=sign_start_row+1, start_column=col_st, end_row=sign_start_row+1, end_column=col_ed)
        c_sig2 = ws.cell(row=sign_start_row+1, column=col_st, value=f"({line2})")
        c_sig2.alignment = Alignment(horizontal="center")
        c_sig2.font = font_bold

        ws.merge_cells(start_row=sign_start_row+2, start_column=col_st, end_row=sign_start_row+2, end_column=col_ed)
        c_sig3 = ws.cell(row=sign_start_row+2, column=col_st, value="วันที่ ......./......./.......")
        c_sig3.alignment = Alignment(horizontal="center")
        c_sig3.font = font_data

    # Save to a temporary or requested file
    export_dir = os.path.join(BASE_DIR, "exports")
    os.makedirs(export_dir, exist_ok=True)
    clean_job = re.sub(r'[^A-Za-z0-9_-]', '_', metadata.get('job_no', 'export'))
    filename = f"รายงานคืนเศษเหล็ก_{clean_job}.xlsx"
    filepath = os.path.join(export_dir, filename)
    wb.save(filepath)
    return filepath, filename

class PEAAppHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=BASE_DIR, **kwargs)

    def do_GET(self):
        url_parts = urllib.parse.urlparse(self.path)
        path = url_parts.path
        query = urllib.parse.parse_qs(url_parts.query)

        if path == "/api/status":
            available = get_available_desktop_files()
            self.send_json_response({
                "status": "ok",
                "default_pdf_exists": len(available) > 0,
                "available_files": available,
                "default_excel_exists": os.path.exists(DEFAULT_EXCEL_PATH),
                "is_vercel": bool(os.environ.get("VERCEL"))
            })
            return

        elif path == "/api/desktop-samples":
            self.send_json_response({
                "files": get_available_desktop_files()
            })
            return

        elif path == "/api/master-data":
            data = load_master_data()
            self.send_json_response(data)
            return

        elif path == "/api/load-desktop-sample":
            raw_target = query.get('file', [''])[0]
            try:
                target_filename = urllib.parse.unquote(raw_target).encode('latin1').decode('utf-8').strip()
            except Exception:
                target_filename = urllib.parse.unquote(raw_target).strip()

            available = get_available_desktop_files()
            target_path = None
            display_name = "018.pdf"

            if target_filename:
                for af in available:
                    if af['filename'] == target_filename or target_filename in af['filename'] or os.path.basename(af['path']) == target_filename:
                        target_path = af['path']
                        display_name = af['filename']
                        break

            if not target_path:
                if os.path.exists(DEFAULT_PDF_PATH):
                    target_path = DEFAULT_PDF_PATH
                    display_name = os.path.basename(DEFAULT_PDF_PATH)
                elif available:
                    target_path = available[0]['path']
                    display_name = available[0]['filename']

            if target_path and os.path.exists(target_path):
                try:
                    if target_path.lower().endswith('.xlsx') or target_path.lower().endswith('.xls'):
                        result = parse_sap_excel(target_path)
                    else:
                        result = parse_sap_pdf(target_path)
                    master = load_master_data()
                    enriched = match_items_with_master(result['items'], master)
                    self.send_json_response({
                        "success": True,
                        "filename": display_name,
                        "metadata": result['metadata'],
                        "items": enriched
                    })
                except ValueError as ve:
                    self.send_json_response({"success": False, "error": str(ve)}, status=400)
                except Exception as e:
                    self.send_json_response({"success": False, "error": str(e)}, status=500)
            else:
                self.send_json_response({"success": False, "error": f"ไม่พบไฟล์ {display_name} บน Desktop"}, status=404)
            return

        elif path.startswith("/exports/"):
            # Serve exported files
            super().do_GET()
            return

        return super().do_GET()

    def do_POST(self):
        url_parts = urllib.parse.urlparse(self.path)
        path = url_parts.path

        content_length = int(self.headers.get('Content-Length', 0))

        if path in ["/api/parse-pdf", "/api/parse-file"]:
            # Handle PDF or Excel upload
            content_type = self.headers.get('Content-Type', '')
            if 'multipart/form-data' in content_type:
                body = self.rfile.read(content_length)

                # Extract boundary
                m_b = re.search(r'boundary=([^;]+)', content_type)
                if m_b:
                    b_str = m_b.group(1).strip('"\'').strip()
                else:
                    b_str = content_type.split("boundary=")[1].strip()

                boundary = b_str.encode('latin1')
                # Multipart segments delimiter is --boundary
                parts = body.split(b"--" + boundary)
                file_bytes = None
                filename = "uploaded_file"

                for part in parts:
                    if b'filename=' in part:
                        # Extract filename
                        header_part = part[:1000].decode('utf-8', errors='ignore')
                        m_fn = re.search(r'filename="?([^";\r\n]+)"?', header_part)
                        if m_fn:
                            filename = os.path.basename(m_fn.group(1).strip())
                        # Extract binary data after double CRLF
                        header_end = part.find(b"\r\n\r\n")
                        if header_end != -1:
                            raw_data = part[header_end+4:]
                            if raw_data.endswith(b"\r\n"):
                                raw_data = raw_data[:-2]
                            file_bytes = raw_data
                            break

                if file_bytes:
                    try:
                        # Detect if Excel or PDF
                        is_excel = (
                            filename.lower().endswith('.xlsx')
                            or filename.lower().endswith('.xls')
                            or file_bytes.startswith(b'PK\x03\x04')
                        )

                        if is_excel:
                            result = parse_sap_excel(file_bytes)
                        else:
                            result = parse_sap_pdf(file_bytes)

                        master = load_master_data()
                        enriched = match_items_with_master(result['items'], master)
                        self.send_json_response({
                            "success": True,
                            "filename": filename,
                            "metadata": result['metadata'],
                            "items": enriched
                        })
                    except ValueError as ve:
                        self.send_json_response({"success": False, "error": str(ve)}, status=400)
                    except Exception as e:
                        self.send_json_response({"success": False, "error": f"เกิดข้อผิดพลาดในการอ่านไฟล์: {str(e)}"}, status=500)
                else:
                    self.send_json_response({"success": False, "error": "ไม่พบข้อมูลไฟล์ในคำขออัปโหลด"}, status=400)
            else:
                self.send_json_response({"success": False, "error": "รูปแบบคำขอต้องเป็น multipart/form-data"}, status=400)
            return

        elif path == "/api/export-excel":
            body = self.rfile.read(content_length)
            try:
                data = json.loads(body.decode('utf-8'))
                filepath, filename = generate_export_excel(data)
                download_url = f"/exports/{urllib.parse.quote(filename)}"
                self.send_json_response({
                    "success": True,
                    "filename": filename,
                    "download_url": download_url
                })
            except Exception as e:
                self.send_json_response({"success": False, "error": f"Export error: {str(e)}"}, status=500)
            return

        elif path == "/api/save-master-data":
            body = self.rfile.read(content_length)
            try:
                items = json.loads(body.decode('utf-8'))
                with open(MASTER_JSON_PATH, "w", encoding="utf-8") as f:
                    json.dump(items, f, ensure_ascii=False, indent=2)
                self.send_json_response({"success": True, "count": len(items)})
            except Exception as e:
                self.send_json_response({"success": False, "error": str(e)}, status=500)
            return

        self.send_json_response({"error": "Endpoint not found"}, status=404)

    def send_json_response(self, data, status=200):
        self.send_response(status)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
        self.wfile.write(json.dumps(data, ensure_ascii=False).encode('utf-8'))

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

def run_server():
    from http.server import ThreadingHTTPServer
    server_address = ('', PORT)
    httpd = ThreadingHTTPServer(server_address, PEAAppHandler)
    print(f"PEA Scrap Calculator server running at http://localhost:{PORT}")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\nShutting down server...")
        httpd.server_close()

if __name__ == "__main__":
    run_server()
